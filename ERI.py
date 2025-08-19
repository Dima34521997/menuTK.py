import json
import os
from copy import copy

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font

# Мои импорты
import obshiy_perechen
import editors as ed

# Генератор ПЭ

cat_names_plural = ed.cat_names_plural
cat_names_singular = ed.cat_names_singular
dfs = []  # DataFrame'ы с изначальными данными
dict_chars = {}  # Словарь комбинированных изначальных DataFrame'ов с их буквенными обозначениями
final_df = []  # Список финальных DataFrame'ов на вставку в документ(-ы)
files = []  # Массив названий файлов для обработки
names_df = {}  # Словарь с формированными строками для вставки в документы
input_files_folder: str  # Путь к файлам
prim_cats = {}  # Категории примечаний
prim_not_install = {}  # Категории для "Не устанавливать"

prog_dir = os.path.dirname(os.path.abspath(__file__))

data: json
coeff: int


def create_names_for_ERI(d_chars: dict):
    """
    Создает лист с компонентами, разделенных по категориям, в котором поля пересобраны по столбцам документа

    :param d_chars: Словарь с сортированными по обозначениям компонентами
    :return: Готовый лист с перечнем компонентов для вставки в документ
    """
    pos_num = 2
    for char in sorted(d_chars.keys()):
        names_df[char] = []
        d_chars[char].fillna('', inplace=True)

        i = 0
        c = 0
        while c < len(d_chars[char]):
            q = d_chars[char]['Quantity'][c]
            module = d_chars[char]['Module'][c]

            name, name2, name3 = ed.create_name(d_chars[char], c, 100)

            # Формирование полей на вставку
            names_df[char].append(pd.Series(dtype=object))
            names_df[char][i]['Поз.'] = pos_num
            names_df[char][i]['Наименование'] = name
            names_df[char][i]['Кол-во на закупку'] = q * coeff
            names_df[char][i]['На 1 прибор'] = q
            names_df[char][i]['Куда входит'] = module
            names_df[char][i]['Примечание'] = d_chars[char]['Примечание'][c]

            if name2 != '':
                i += 1
                names_df[char].append(pd.Series())
                names_df[char][i]['Поз.'] = ''
                names_df[char][i]['Наименование'] = name2
                names_df[char][i]['Кол-во на закупку'] = ''
                names_df[char][i]['На 1 прибор'] = ''
                names_df[char][i]['Куда входит'] = ''
                names_df[char][i]['Примечание'] = ''

            if name3 != '':
                i += 1
                names_df[char].append(pd.Series())
                names_df[char][i]['Поз.'] = ''
                names_df[char][i]['Наименование'] = name3
                names_df[char][i]['Кол-во на закупку'] = ''
                names_df[char][i]['На 1 прибор'] = ''
                names_df[char][i]['Куда входит'] = ''
                names_df[char][i]['Примечание'] = ''

            pos_num += 1
            i += 1
            c += 1
    return names_df


def export_to_excel(names_df):
    global coeff

    module = pd.read_excel(files[0], sheet_name='BOM', header=None).loc[7, 10].split(' ')[0]
    save_dir = files[0][:-len(files[0].split("\\")[-1])] + "\\ЭРИ\\"
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)

    wb = load_workbook("Шаблоны/ЭРИ.xlsx")
    sheet = wb.worksheets[0]

    paste_index = 6

    sheet.cell(paste_index, 2).value = module
    paste_index += 1

    for char in sorted(dict_chars.keys()):
        for index, chip in enumerate(names_df[char]):
            if index == 0:
                # Поиск наименования категории
                cat_name = 'Нет категории'
                if len(names_df[char]) > 1:
                    for desig, d_cat_name in cat_names_plural.items():
                        if char == desig:
                            cat_name = d_cat_name
                            break
                else:
                    for desig, d_cat_name in cat_names_singular.items():
                        if char == desig:
                            cat_name = d_cat_name
                            break

                ft = Font(name='Book Antiqua',
                          size=10,
                          bold=True)
                alignment = Alignment(horizontal="center", vertical="center")

                sheet.insert_rows(paste_index)

                sheet.cell(paste_index, 2).font = ft
                sheet.cell(paste_index, 2).alignment = alignment
                sheet.cell(paste_index, 2).fill = PatternFill("solid", fgColor="0099CC00")
                sheet.cell(paste_index, 2).value = cat_name

                paste_index += 1

            sheet.insert_rows(paste_index)
            font = copy(sheet.cell(6, 2).font)
            borders = copy(sheet.cell(6, 2).border)
            alignment = copy(sheet.cell(6, 2).alignment)

            for cell in sheet[paste_index]:
                cell.font = font
                cell.border = borders
                cell.alignment = alignment

            sheet.cell(paste_index, 1).value = chip['Поз.']
            sheet.cell(paste_index, 1).alignment = Alignment(horizontal='center', vertical='center')

            sheet.cell(paste_index, 2).value = chip['Наименование']

            sheet.cell(paste_index, 3).value = chip['Кол-во на закупку']
            sheet.cell(paste_index, 3).alignment = Alignment(horizontal='center', vertical='center')

            sheet.cell(paste_index, 4).value = chip['На 1 прибор']
            sheet.cell(paste_index, 4).alignment = Alignment(horizontal='center', vertical='center')

            sheet.cell(paste_index, 5).value = chip['Куда входит']

            paste_index += 1

    sheet.cell(1, 6).value = coeff
    sheet.cell(1, 6).font = font
    sheet.cell(1, 6).alignment = Alignment(horizontal='center', vertical='center')
    new_name = module
    wb.save(save_dir + new_name + " ЭРИ.xlsx")


def execute(input_files):
    global dfs, files, dict_chars, coeff
    files = input_files
    dfs = []
    d_chars = {}

    obshiy_perechen.test = False

    with open("Profile.json", "r") as read_file:
        data = json.load(read_file)

    obshiy_perechen.no_perechen = 0
    print("Получаю данные из перечней элементов...")
    dfs, files = obshiy_perechen.get_dfs(d_chars, files)
    print("Формирую общую таблицу элементов...")
    d_chars, prim_not_install, prim_cats, one_man_cats = obshiy_perechen.get_components(d_chars, dfs, files)
    print("Комбинирую компоненты...")
    d_chars = obshiy_perechen.combine_chips_in_module(d_chars)

    coeff = data["SB_count"]

    print("Формирую итоговый перечень для вставки...")
    names_df = create_names_for_ERI(d_chars)
    print("Вставляю готовый перечень в шаблон...")
    export_to_excel(names_df)
    print("Готово!")
